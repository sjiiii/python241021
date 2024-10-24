import openpyxl
from openpyxl.styles import Font, Alignment
import random

# 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "전자제품 판매 데이터"

# 헤더 추가
headers = ["제품ID", "제품명", "수량", "가격"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 제품 목록
products = [
    "스마트폰", "노트북", "태블릿", "스마트워치", "무선이어폰", 
    "블루투스 스피커", "게이밍 마우스", "기계식 키보드", "모니터", "프린터"
]

# 데이터 생성 및 추가
for row in range(2, 102):  # 100개의 데이터 생성
    product_id = row - 1
    product_name = random.choice(products)
    quantity = random.randint(1, 100)
    price = random.randint(10000, 2000000)  # 1만원 ~ 200만원

    ws.cell(row=row, column=1, value=product_id)
    ws.cell(row=row, column=2, value=product_name)
    ws.cell(row=row, column=3, value=quantity)
    ws.cell(row=row, column=4, value=price)

# 열 너비 조정
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# 엑셀 파일 저장
wb.save("products.xlsx")

print("products.xlsx 파일이 생성되었습니다.")